# GIC 학회 표준 5개년 모델 Excel 빌드 (v10.0 트랙 2)
# 출력: templates/Excel_5개년모델_v10.xlsx (8 sheets)

from __future__ import annotations
import io
import sys
from pathlib import Path

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8")

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────
# 학회 표준 토큰
# ─────────────────────────────────────────────────────────────
COLOR = {
    "input": "FFE9D6",       # 오렌지 10%
    "formula": "FFFFFF",
    "reference": "E8EFFF",   # 네이비 10%
    "header_bg": "072A51",
    "header_fg": "FFFFFF",
    "ok_bg": "22C55E",
    "err_bg": "EF4444",
}

BORDER = Border(
    left=Side(style="thin", color="EAEAEA"),
    right=Side(style="thin", color="EAEAEA"),
    top=Side(style="thin", color="EAEAEA"),
    bottom=Side(style="thin", color="EAEAEA"),
)


def fill(hex_):
    return PatternFill("solid", fgColor=hex_)


def header_cell(ws, cell_addr, text):
    ws[cell_addr] = text
    ws[cell_addr].fill = fill(COLOR["header_bg"])
    ws[cell_addr].font = Font(bold=True, color=COLOR["header_fg"], size=11)
    ws[cell_addr].alignment = Alignment(horizontal="center", vertical="center")
    ws[cell_addr].border = BORDER


def input_cell(ws, cell_addr, value=None):
    if value is not None:
        ws[cell_addr] = value
    ws[cell_addr].fill = fill(COLOR["input"])
    ws[cell_addr].border = BORDER
    ws[cell_addr].alignment = Alignment(horizontal="right")


def formula_cell(ws, cell_addr, formula):
    ws[cell_addr] = formula
    ws[cell_addr].fill = fill(COLOR["formula"])
    ws[cell_addr].border = BORDER
    ws[cell_addr].alignment = Alignment(horizontal="right")


def label_cell(ws, cell_addr, text):
    ws[cell_addr] = text
    ws[cell_addr].border = BORDER
    ws[cell_addr].font = Font(bold=False, size=11)


# ─────────────────────────────────────────────────────────────
# 시트 빌더
# ─────────────────────────────────────────────────────────────
def build_assumptions(wb):
    ws = wb.create_sheet("Assumptions")
    ws.sheet_properties.tabColor = "072A51"

    ws["A1"] = "1. Assumptions — Bear / Base / Bull 가정표"
    ws["A1"].font = Font(bold=True, size=14, color="072A51")
    ws.merge_cells("A1:D1")

    ws["A2"] = "기업명·분석 기준일: [입력]"
    ws["A2"].fill = fill(COLOR["input"])

    headers = ["항목", "Bear", "Base", "Bull"]
    for i, h in enumerate(headers):
        header_cell(ws, f"{get_column_letter(i+1)}3", h)

    rows = [
        ("매출 5Y CAGR (%)", -2.0, 8.0, 18.0),
        ("영업이익률 FY5 (%)", 8.0, 14.0, 18.0),
        ("적용 PER (배)", 12.0, 18.0, 24.0),
        ("적용 EV/EBITDA (배)", 6.0, 9.0, 12.0),
        ("적용 PBR (배)", 1.0, 1.5, 2.2),
        ("WACC (%)", 9.0, 8.0, 7.0),
        ("Terminal Growth (%)", 1.0, 2.0, 3.0),
    ]
    for r, (label, bear, base, bull) in enumerate(rows, start=4):
        label_cell(ws, f"A{r}", label)
        input_cell(ws, f"B{r}", bear)
        input_cell(ws, f"C{r}", base)
        input_cell(ws, f"D{r}", bull)

    ws.column_dimensions["A"].width = 30
    for col in ["B", "C", "D"]:
        ws.column_dimensions[col].width = 14


def build_is(wb):
    ws = wb.create_sheet("IS")
    ws.sheet_properties.tabColor = "0D4889"

    ws["A1"] = "2. IS — 5개년 손익계산서 (단위: 십억원)"
    ws["A1"].font = Font(bold=True, size=14, color="072A51")
    ws.merge_cells("A1:G1")

    headers = ["항목", "FY1", "FY2", "FY3E", "FY4E", "FY5E", "비고"]
    for i, h in enumerate(headers):
        header_cell(ws, f"{get_column_letter(i+1)}3", h)

    items = [
        ("매출액", "input"),
        ("매출원가", "input"),
        ("매출총이익", "formula:=B4-B5"),  # 첫 컬럼만 표시, 다른 컬럼은 자동
        ("매출총이익률 (%)", "formula:=IFERROR(B6/B4*100,0)"),
        ("판관비", "input"),
        ("영업이익", "formula:=B6-B8"),
        ("영업이익률 (%)", "formula:=IFERROR(B9/B4*100,0)"),
        ("당기순이익", "input"),
        ("순이익률 (%)", "formula:=IFERROR(B11/B4*100,0)"),
        ("매출 YoY (%)", "formula:=IFERROR((B4/A4-1)*100,0)"),
    ]

    for r_offset, (label, kind) in enumerate(items):
        row = 4 + r_offset
        label_cell(ws, f"A{row}", label)
        for col_idx in range(2, 7):  # B~F (FY1~FY5)
            col = get_column_letter(col_idx)
            cell = f"{col}{row}"
            if kind == "input":
                input_cell(ws, cell)
            else:
                # formula를 컬럼별로 적응 (첫 컬럼 기준 식을 다른 컬럼으로 평행이동)
                base_formula = kind.split(":", 1)[1]
                # 단순한 컬럼 변환 (B → 현재 컬럼)
                shifted = base_formula.replace("B", col).replace("A", get_column_letter(col_idx - 1))
                if col_idx == 2 and "/A4" in base_formula:
                    # FY1의 YoY는 데이터 없음
                    formula_cell(ws, cell, "")
                else:
                    formula_cell(ws, cell, shifted)

    # 비고 컬럼 (G)
    notes = ["", "", "(자동 계산)", "(자동 계산)", "", "(자동 계산)", "(자동 계산)", "", "(자동 계산)", "(자동 계산)"]
    for r_offset, note in enumerate(notes):
        if note:
            ws[f"G{4 + r_offset}"] = note
            ws[f"G{4 + r_offset}"].font = Font(italic=True, color="8A8A8A", size=10)

    ws.column_dimensions["A"].width = 22
    for col in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 12
    ws.column_dimensions["G"].width = 16


def build_bs(wb):
    ws = wb.create_sheet("BS")
    ws.sheet_properties.tabColor = "0D4889"

    ws["A1"] = "3. BS — 5개년 재무상태표 (단위: 십억원)"
    ws["A1"].font = Font(bold=True, size=14, color="072A51")
    ws.merge_cells("A1:G1")

    headers = ["항목", "FY1", "FY2", "FY3E", "FY4E", "FY5E", "Balance Check"]
    for i, h in enumerate(headers):
        header_cell(ws, f"{get_column_letter(i+1)}3", h)

    items_input = ["자산총계", "부채총계", "자본총계", "현금성자산", "차입금"]
    for r_offset, label in enumerate(items_input):
        row = 4 + r_offset
        label_cell(ws, f"A{row}", label)
        for col_idx in range(2, 7):
            col = get_column_letter(col_idx)
            input_cell(ws, f"{col}{row}")

    # Balance Check 행
    bcr = 4 + len(items_input)
    label_cell(ws, f"A{bcr}", "Balance Check")
    for col_idx in range(2, 7):
        col = get_column_letter(col_idx)
        formula_cell(ws, f"{col}{bcr}", f'=IF({col}4={col}5+{col}6,"OK","ERROR")')
        # 조건부 서식: OK 초록, ERROR 빨강
        ws.conditional_formatting.add(
            f"{col}{bcr}",
            CellIsRule(operator="equal", formula=['"OK"'], fill=fill(COLOR["ok_bg"]), font=Font(color="FFFFFF", bold=True)),
        )
        ws.conditional_formatting.add(
            f"{col}{bcr}",
            CellIsRule(operator="equal", formula=['"ERROR"'], fill=fill(COLOR["err_bg"]), font=Font(color="FFFFFF", bold=True)),
        )

    # 순차입금 행
    nd_row = bcr + 1
    label_cell(ws, f"A{nd_row}", "순차입금 (= 차입금 - 현금성자산)")
    for col_idx in range(2, 7):
        col = get_column_letter(col_idx)
        formula_cell(ws, f"{col}{nd_row}", f"={col}8-{col}7")

    ws.column_dimensions["A"].width = 28
    for col in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 13


def build_cf(wb):
    ws = wb.create_sheet("CF")
    ws.sheet_properties.tabColor = "0D4889"

    ws["A1"] = "4. CF — 5개년 현금흐름표 (단위: 십억원)"
    ws["A1"].font = Font(bold=True, size=14, color="072A51")
    ws.merge_cells("A1:G1")

    headers = ["항목", "FY1", "FY2", "FY3E", "FY4E", "FY5E", "비고"]
    for i, h in enumerate(headers):
        header_cell(ws, f"{get_column_letter(i+1)}3", h)

    inputs = ["영업CF", "투자CF", "재무CF", "CAPEX"]
    for r_offset, label in enumerate(inputs):
        row = 4 + r_offset
        label_cell(ws, f"A{row}", label)
        for col_idx in range(2, 7):
            col = get_column_letter(col_idx)
            input_cell(ws, f"{col}{row}")

    # FCF 행
    fcf_row = 4 + len(inputs)
    label_cell(ws, f"A{fcf_row}", "FCF (= 영업CF - CAPEX)")
    for col_idx in range(2, 7):
        col = get_column_letter(col_idx)
        formula_cell(ws, f"{col}{fcf_row}", f"={col}4-{col}7")

    # CAPEX/매출
    cap_row = fcf_row + 1
    label_cell(ws, f"A{cap_row}", "CAPEX / 매출 (%)")
    for col_idx in range(2, 7):
        col = get_column_letter(col_idx)
        formula_cell(ws, f"{col}{cap_row}", f"=IFERROR({col}7/IS!{col}4*100,0)")

    ws.column_dimensions["A"].width = 28
    for col in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 12


def build_valuation(wb):
    ws = wb.create_sheet("Valuation")
    ws.sheet_properties.tabColor = "E57728"

    ws["A1"] = "5. Valuation — 4종 밸류에이션 (단위: 원)"
    ws["A1"].font = Font(bold=True, size=14, color="072A51")
    ws.merge_cells("A1:F1")

    headers = ["방법론", "Bear", "Base", "Bull", "가중 평균", "비고"]
    for i, h in enumerate(headers):
        header_cell(ws, f"{get_column_letter(i+1)}3", h)

    methods = [
        ("PER 적용", "FY5 EPS × 적용 PER"),
        ("EV/EBITDA 적용", "(EBITDA × 멀티플 - 순차입금) / 주식수"),
        ("PBR 적용", "BPS × 적용 PBR"),
        ("DCF (선택)", "5개년 FCF 할인 + TV"),
        ("최종 목표주가", "= AVERAGE(B7:D7)"),
    ]
    for r_offset, (m, note) in enumerate(methods):
        row = 4 + r_offset
        label_cell(ws, f"A{row}", m)
        for col in ["B", "C", "D"]:
            input_cell(ws, f"{col}{row}")
        formula_cell(ws, f"E{row}", f"=AVERAGE(B{row}:D{row})")
        ws[f"F{row}"] = note
        ws[f"F{row}"].font = Font(italic=True, color="8A8A8A", size=10)

    ws.column_dimensions["A"].width = 22
    for col in ["B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["F"].width = 30


def build_scenario(wb):
    ws = wb.create_sheet("Scenario")
    ws["A1"] = "6. Scenario — EPS × PER 매트릭스"
    ws["A1"].font = Font(bold=True, size=14, color="072A51")
    ws.merge_cells("A1:F1")

    ws["A3"] = "EPS \\ PER"
    ws["A3"].fill = fill(COLOR["header_bg"])
    ws["A3"].font = Font(bold=True, color="FFFFFF")
    pers = [16, 18, 20, 22, 24]
    for i, p in enumerate(pers):
        cell = f"{get_column_letter(2 + i)}3"
        ws[cell] = f"PER {p}x"
        ws[cell].fill = fill(COLOR["header_bg"])
        ws[cell].font = Font(bold=True, color="FFFFFF")

    eps_labels = ["EPS Bear", "EPS Base", "EPS Bull"]
    for r_offset, label in enumerate(eps_labels):
        row = 4 + r_offset
        label_cell(ws, f"A{row}", label)
        input_cell(ws, f"B{row}")  # EPS 값 입력
        # B 컬럼은 EPS 값. C~F에 EPS × PER 자동
        for i, p in enumerate(pers):
            col = get_column_letter(2 + i)
            formula_cell(ws, f"{col}{row}", f"=$B${row}*{p}")

    ws.column_dimensions["A"].width = 14


def build_sensitivity(wb):
    ws = wb.create_sheet("Sensitivity")
    ws["A1"] = "7. Sensitivity — 매출 ±15% × OPM ±2%p 민감도"
    ws["A1"].font = Font(bold=True, size=14, color="072A51")
    ws.merge_cells("A1:F1")

    ws["A3"] = "매출 \\ OPM"
    ws["A3"].fill = fill(COLOR["header_bg"])
    ws["A3"].font = Font(bold=True, color="FFFFFF")
    opms = ["-2%p", "-1%p", "Base", "+1%p", "+2%p"]
    for i, o in enumerate(opms):
        cell = f"{get_column_letter(2 + i)}3"
        ws[cell] = o
        ws[cell].fill = fill(COLOR["header_bg"])
        ws[cell].font = Font(bold=True, color="FFFFFF")

    revenue_scenarios = ["-15%", "-7.5%", "Base", "+7.5%", "+15%"]
    for r_offset, label in enumerate(revenue_scenarios):
        row = 4 + r_offset
        label_cell(ws, f"A{row}", label)
        for i in range(5):
            col = get_column_letter(2 + i)
            input_cell(ws, f"{col}{row}")

    ws.column_dimensions["A"].width = 14


def build_summary(wb):
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = "E57728"

    ws["A1"] = "8. Summary — 1페이지 요약 (다른 시트 참조)"
    ws["A1"].font = Font(bold=True, size=14, color="072A51")
    ws.merge_cells("A1:D1")

    rows = [
        ("종목명·종목코드", "[Sheet 1 입력 값 자동 참조]"),
        ("분석 기준일", "[Sheet 1 입력]"),
        ("투자의견", "[BUY/HOLD/SELL]"),
        ("목표주가 (가중)", "=Valuation!E8"),
        ("현재가", "[입력]"),
        ("상승여력 (%)", "=IFERROR((B6-B7)/B7*100,0)"),
        ("FY1 매출 (십억원)", "=IS!B4"),
        ("FY1 영업이익", "=IS!B9"),
        ("FY1 순이익", "=IS!B11"),
        ("FY1 ROE (%)", "[Sheet 입력 또는 자동]"),
        ("Balance Check (FY1)", "=BS!B9"),
    ]
    for r_offset, (label, ref) in enumerate(rows):
        row = 3 + r_offset
        label_cell(ws, f"A{row}", label)
        if ref.startswith("="):
            formula_cell(ws, f"B{row}", ref)
        elif ref.startswith("[") and ref.endswith("]"):
            input_cell(ws, f"B{row}")
        else:
            ws[f"B{row}"] = ref

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 35


def build_template(output_path: Path):
    wb = Workbook()
    # 기본 시트 제거
    wb.remove(wb.active)

    build_assumptions(wb)
    build_is(wb)
    build_bs(wb)
    build_cf(wb)
    build_valuation(wb)
    build_scenario(wb)
    build_sensitivity(wb)
    build_summary(wb)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"OK: {output_path}")
    print(f"sheets: {wb.sheetnames}")


def main():
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--output", default=None)
    args = ap.parse_args()

    if args.output:
        out = Path(args.output)
    else:
        script = Path(__file__).resolve()
        v10 = script.parents[4]
        out = v10 / "templates" / "Excel_5개년모델_v10.xlsx"
    build_template(out)


if __name__ == "__main__":
    main()
